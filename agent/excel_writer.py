import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def _safe_filename(text: str) -> str:
    text = text.lower().replace(" ", "_")
    return re.sub(r"[^a-z0-9_]", "", text)


def save_leads_to_excel(leads: list[dict], category: str, location: str) -> str:
    filename = f"leads_{_safe_filename(category)}_{_safe_filename(location)}.xlsx"
    filepath = str(Path.cwd() / filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    headers = ["Business Name", "Email", "Phone Number", "Website", "Location"]
    header_font = Font(bold=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font

    for row_idx, lead in enumerate(leads, 2):
        ws.cell(row=row_idx, column=1, value=lead.get("name", ""))
        ws.cell(row=row_idx, column=2, value=lead.get("email", ""))
        ws.cell(row=row_idx, column=3, value=lead.get("phone", ""))
        ws.cell(row=row_idx, column=4, value=lead.get("website", ""))
        ws.cell(row=row_idx, column=5, value=location)

    for col_idx in range(1, 6):
        lengths = []
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            val = row[0]
            if val is not None:
                lengths.append(len(str(val)))
        max_len = max(lengths) if lengths else 0
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    wb.save(filepath)
    return filepath
