from openpyxl import load_workbook
from agent.excel_writer import save_leads_to_excel

fake_leads = [
    {"name": "Cafe Blue", "phone": "+921234567890", "website": "https://cafeblue.pk", "email": ""},
    {"name": "Green Cup Coffee", "phone": "", "website": "https://greencup.pk"},
    {"name": "Roast House", "phone": "+929876543210", "website": ""},
]

filepath = save_leads_to_excel(fake_leads, "coffee shops", "Karachi")
print(f"Saved to: {filepath}")

wb = load_workbook(filepath)
ws = wb.active
for row in ws.iter_rows(min_row=1, values_only=True):
    print(list(row))
