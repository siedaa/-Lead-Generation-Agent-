import glob
import os

from openpyxl import load_workbook

xlsx_files = glob.glob("leads_*.xlsx")
if not xlsx_files:
    print("No leads_*.xlsx files found.")
else:
    latest = max(xlsx_files, key=os.path.getmtime)
    print(f"Opening most recent file: {latest}\n")
    wb = load_workbook(latest)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        print(list(row))